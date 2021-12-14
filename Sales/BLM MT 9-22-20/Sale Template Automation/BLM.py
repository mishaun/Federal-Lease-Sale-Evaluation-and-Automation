#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Sun Feb  9 20:57:29 2020

@author: Mishaun_Bhakta

The purpose of this program is to automate the process of preparing and closing 
Bureau of Land Management (BLM) federal oil and gas lease sales

The first phase of the program will visit the website, scrape the lease information
for each tract, and parse the information into a template sale note spreadsheet
It will also download the shapefile of the sale, which is needed for evaluating tracts on
DrillingInfo.com

Once the spreadsheet is prepared and the sale takes place, the program will 
scrape the webpage again and gather information regarding won parcels.
The scrape will get the bonus bid for each parcel we won by evaluating our bidder number

Finally, a dataframe will be created based on won lots and the information will be 
passed into a pdf fill form function to create paperwork needed to send to 
Bureau of Land Management

"""
import os, re, shutil, time

#sale parameters
state = "Montana"
stinitials = "MT"
date = "Sep 22, 2020"
bidder = '3'
url = 'https://www.energynet.com/govt_listing.pl?sg=5314'
bidDuration = 1 #amount of hours the parcel is open for bidding

#Navigate to energynet/govt sale and get sale page
from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup

#getting filepath of this file
filepath = os.path.dirname(__file__)

#driver will be used based on operating system - windows or mac
try:
    driver = webdriver.Chrome(filepath + "/chromedriver.exe")
except:
    driver = webdriver.Chrome(filepath + "/chromedriver")

#selenium driver object will go to url of sale
driver.implicitly_wait(30)
driver.get(url)

#finding the show dropdown menu
showButton = driver.find_elements_by_css_selector('#DataTables_Table_0_length > label > select')[0]
                                                   
#sending keys to select all lots
showButton.send_keys("All")
showButton.send_keys(Keys.RETURN)
                                                 

#storing html content in variable after reaching target sale page
salehtml = BeautifulSoup(driver.page_source, "html.parser")

def webscrape_presale(parsepage):
    '''This function will take a page and scrape its data for sale lot information
    It will also download sale shapefile and move it to directory of this script file
    '''

    #webscrape sale page - gathering lot serial numbers from html 
    serialnums = parsepage.find_all("span", "lot-name")
    #ist comprehension - appending text into serialnums
    serialnums = [i.text for i in serialnums]
    
    #storing all data from tag 'td's with clas name "lot-legal
    #this html container/tag has 3 pieces of information
    legalinfo = parsepage.find_all("td", "lot-legal")
    
    lotclosings = parsepage.find_all("td", "lot-closing")
    
    
    closingInfo = []
    for item in lotclosings:
        #finding date from lot closing element
        closingDate = re.search("\d+/\d+/\d+", item.text)[0]
        
        #finding opening time from lot closing element - then calculate closing time
        openingTime = re.search("\d+:\d+", item.text)[0]
        #calculating closing time 
        closingTime = str(int(openingTime.split(":")[0]) + bidDuration)+":" + openingTime.split(":")[1]
        
        dateAndTime = closingDate + " " + closingTime
        closingInfo.append(dateAndTime)
    
    #initializing empty arrays
    acres = []
    desc = []
    county = []
    
    for item in legalinfo:
        county.append(item.contents[0].text)
        desc.append(item.contents[1].text)
        #getting acres by splitting at : and blankspace to get string of numerical value - taking out a comma if above 1000 in order to convert to float
        acres.append(float(re.split(":\W",item.contents[2].text)[1].replace(',','')))

    ##getting shapefile from webpage  
    #clicking link of where shapefile is stored on sale page    
    driver.find_element_by_link_text("GIS Data WGS84").click()
    time.sleep(2)
    try:
        driver.find_element_by_link_text("Notice of Competitive Oil and Gas Internet-Based Lease Sale").click()
    except:
        print("Sale notice pdf was unable to be clicked")
    #getting list of filenames in downloads
    try:
        downloaddir = "/Users/Mishaun_Bhakta/Downloads/"
        downloads = os.listdir(downloaddir)
    except:
        downloaddir = "C:/Users/mishaun/Downloads/"
        downloads = os.listdir(downloaddir) 
        
    #pattern will find downloaded file name of shapefile
    pattern = "BLM"+ stinitials + "\S*.zip"
    
    #searching through filenames in downlaods folder
    finds = []
    for file in downloads:
        if re.findall(pattern, file):
            finds.append(file)
            break
        
    #moving file from downloads folder to directory of this script file - then renaming it to a cleaner name
    shutil.copy(downloaddir + finds[0], filepath)
    
    try:
        os.rename(finds[0], "BLM " + stinitials + " " + date + " Shapefile." + finds[0].split(".")[1])
    except:
       pass
   
    return acres, desc, county, serialnums, closingInfo

#storing global variables of web scrape information
acres, descriptions, counties, serials, closing = webscrape_presale(salehtml)

#Open sale template and update insert information from webscrape
import openpyxl

def fillexcel():
    '''
    This function will take scraped (global) values for lots and insert into sale spreadsheet 
    '''    
    
    #opening template sale notebook for modifications
    #preserving vba to keep formatting of workbook preserved - also keeping formulas 
    wb = openpyxl.load_workbook("BLM Sale Notes Template.xlsm", keep_vba = True)
    sheet = wb.active
    
    #updating sheet title to sale title
    sheet["B6"] = "BLM {} {} Sale Notes".format(stinitials, date)
    
    #inserting values from webscrape into spreadsheet -8th row is where data rows begin
    for i in range(0,len(serials)):
        sheet.cell(row = 8+i, column = 2, value = serials[i])
        sheet.cell(row = 8+i, column = 5, value = closing[i])
        sheet.cell(row = 8+i, column = 6, value = acres[i])
        sheet.cell(row = 8+i, column = 7, value = counties[i])
        sheet.cell(row = 8+i, column = 8, value = descriptions[i])
    
    #checking to see whether or not excel file already exists - if it does it'll prevent overwriting of changes
    if os.path.exists(filepath+ "/" + "BLM {} {} Sale Notes.xlsm".format(stinitials, date)):
        print("File already exists - Preventing overwrite of changes in excel file")
    else:
        wb.save("BLM {} {} Sale Notes.xlsm".format(stinitials, date))
        wb.close()

#checking to see whether or not excel file already exists - if it does it'll prevent overwriting of changes
if os.path.exists(filepath+ "/" + "BLM {} {} Sale Notes.xlsm".format(stinitials, date)):
    print("File already exists - Preventing overwrite of changes in excel file")
else:
    fillexcel()

bidtags = salehtml.find_all("td", "lot-bid")
ourwinnings = {}

for i in range (0,len(bidtags)):
    textCont = bidtags[i].text
    
    #extracting bidder number from lot bid tags - try statement prevents break if no bids were received
    try:
        winBidder = re.findall('#\d+', textCont)[0].replace("#",'')
    except:
        print("no bids received for parcel: " + serials[i])
        winBidder = type(None)
        pass
    
    #if we won the bid, then capture the winning bid amount
    if bidder == winBidder:
        winAmount = re.findall('\$\d+', textCont)[0]
        winAmount = winAmount.replace('$','')
    
        ourwinnings[i] = winAmount

def fillwinnings():
    '''This function will take ourwinnings dictionary and add values to created spreadsheet
    '''
    
    #### insert our winnings into sale spreadsheet
    wb = openpyxl.load_workbook("BLM {} {} Sale Notes.xlsm".format(stinitials, date), keep_vba = True)
    sheet = wb.active
    
    for i in range(0,len(ourwinnings)):
        #row 8 is the starting row for parcels in teh spreadsheet, inserting data relative to 8th row by adding parcel number of sale
        sheet.cell(row = 8 + list(ourwinnings.keys())[i], column = 17, value = ourwinnings[list(ourwinnings.keys())[i]])
        sheet.cell(row = 8 + list(ourwinnings.keys())[i], column = 16, value = 'Y')
    
    wb.save("BLM {} {} Sale Notes.xlsm".format(stinitials, date))
    wb.close()

#create dataframe for completed sale sheet
import pandas as pd

#use pdf reader to fill in form
# conda install -c conda-forge pdfrw
import pdfrw

#copied code and function from article: https://bostata.com/how-to-populate-fillable-pdfs-with-python/
##############################################################################
ANNOT_KEY = '/Annots'
ANNOT_FIELD_KEY = '/T'
ANNOT_VAL_KEY = '/V'
ANNOT_RECT_KEY = '/Rect'
SUBTYPE_KEY = '/Subtype'
WIDGET_SUBTYPE_KEY = '/Widget'

def write_fillable_pdf(input_pdf_path, output_pdf_path, data_dict):
    '''
    This function will fill in pdf's forms based on a form pdf
    '''
    
    template_pdf = pdfrw.PdfReader(input_pdf_path)
    annotations = template_pdf.pages[0][ANNOT_KEY]
    for annotation in annotations:
        if annotation[SUBTYPE_KEY] == WIDGET_SUBTYPE_KEY:
            if annotation[ANNOT_FIELD_KEY]:
                key = annotation[ANNOT_FIELD_KEY][1:-1]
                if key in data_dict.keys():
                    annotation.update(
                        pdfrw.PdfDict(V='{}'.format(data_dict[key]))
                    )
    pdfrw.PdfWriter().write(output_pdf_path, template_pdf)
##############################################################


def wonlotsDF():
    '''
    This function will create a dataframe of the won lots by reading information
    from completed sale note spreadsheet
    The dataframe will then be used to parse pdf's 
    '''
    #using openpyxl in order to read formulated values from spreadsheet
    # NOTE: have to manually open excel and save sheet for formulated cells to read after filling in values
    data_onlyWB = openpyxl.load_workbook("BLM {} {} Sale Notes.xlsm".format(stinitials, date), data_only = True, keep_vba = True)
    dataSheet = data_onlyWB.active
    
    #covnerting spreadsheet into dataframe
    df = pd.DataFrame(dataSheet.values)
    
    #slicing the dataframe to get only relevant data
    df = df.iloc[6:,1:25]
    #setting columns to first row of dataframe
    df.columns = df.iloc[0]
    #dropping the repeated row with column names
    df = df.drop(index =[6])
    
    #filtering data frame with values only won by magnum
    wonlotsdf = df[df["Magnum Won (Y/N)"] == 'Y']
    return wonlotsdf


def createBidSheets():
    '''
    This function will take a template pdf and generate pdf's based on wonlots dataframe
    '''
    #calling funciton wonlotsDF in ordre for bid sheets to be created 
    wonlotsdf = wonlotsDF()

    templatePDF = 'bidsheet template.pdf'
    
    for i in range(0,len(wonlotsdf.index)):
    
        OutputPath = filepath +"/Bid Sheets/" + wonlotsdf.iloc[i]["Serial numbers"] + " Bid Sheet.pdf"
        
        fields = {
                "State": stinitials,
                "Date of Sale": date,
                'Check Box for Oil and Gas' : "x",
                "Oil and Gas/Parcel No" : wonlotsdf.iloc[i]["Serial numbers"],
                "TOTAL BID FOR Oil and Gas Lease" : wonlotsdf.iloc[i]["Total Bid (Number on BLM Bid Sheet)"],
                "PAYMENT SUBMITTED WITH BID for Oil and Gas" : wonlotsdf.iloc[i]["Min Due"],
                "Print or Type Name of Lessee" : "R&R Royalty, LTD",
                "Address of Lessee": "500 N Shoreline Blvd, Ste 322",
                "City" : "Corpus Christi",
                "State_2": "TX",
                "Zip Code" : "78401"
                }
        
        write_fillable_pdf(templatePDF, OutputPath, fields)

def openDI():
    '''This function will open up DrillingInfo and log user in
    '''
    
    driver = webdriver.Chrome(filepath + "/chromedriver")
    wait = WebDriverWait(driver, 20)
        
    driver.get("https://app.drillinginfo.com/gallery/")
    userfield = driver.find_element_by_name("username")
    passfield = driver.find_element_by_name("password")
    
    userfield.click()
    userfield.send_keys("mbhaktamgm")
    passfield.click()
    passfield.send_keys("itheCwe")
    passfield.send_keys(Keys.RETURN)
    
    myworkspaces = wait.until(EC.element_to_be_clickable((By.XPATH, '//*[@id=\"workspaces-section\"]/div[1]/div[1]')))
    myworkspaces.click()
    
    default_workspace = wait.until(EC.element_to_be_clickable((By.XPATH, "//*[@id=\"workspaces-section\"]/div[3]/di-carousel/section/div[2]/table/tbody/tr[2]/a/span[2]/span")))
    default_workspace.click()
    
#splitting counties in counties variable at the comma+space to formulate string 
#for drilling info paste and filter
    
splitCounties = [item.split(", ") for item in counties]

DIcounties =  []
for item in splitCounties:
    #taking "County" out of word
    temp = item[0].upper().replace("COUNTY", "")
    formattedCounty = temp + "(" + item[1] + ")"
    DIcounties.append(formattedCounty)



        
    
